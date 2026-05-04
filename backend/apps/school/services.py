import csv
import io
import re
import unicodedata
from datetime import date, timedelta
from pathlib import Path

import requests
from django.db.models import Q
from openpyxl import load_workbook
from pypdf import PdfReader

from .models import AcademicCalendar, CalendarEvent, ClassGroup, EventType, NationalHolidayCache, School, Student, Weekday, WeeklyClassSlot

BRASIL_API_HOLIDAYS = "https://brasilapi.com.br/api/feriados/v1/{year}"

SUPPORTED_IMPORT_EXTENSIONS = {".csv", ".xlsx", ".pdf"}

HEADER_ALIASES = {
    "school": {"school", "escola", "school_name", "nome_escola", "unidade", "nome_da_escola"},
    "city": {"city", "cidade", "municipio"},
    "state": {"state", "estado", "uf"},
    "class_group": {"class_group", "turma", "class", "nome_turma", "turma_nome", "tr"},
    "grade_level": {"grade", "grade_level", "serie", "ano", "nivel"},
    "shift": {"shift", "turno", "periodo"},
    "student": {"student", "aluno", "student_name", "nome_aluno", "nome"},
    "enrollment_code": {"enrollment", "matricula", "enrollment_code", "codigo_matricula", "ra"},
}


def _normalize_key(value: str) -> str:
    # Unifica variações como "nome_aluno" e "nome aluno" no mesmo formato.
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def _normalize_text(value: str) -> str:
    text = str(value or "").strip().lower()
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def _extract_school_and_shift_hint(text: str) -> tuple[str, str]:
    normalized = _normalize_text(text)
    shift_pattern = r"\b(matutino|vespertino|noturno|manha|tarde|noite|morning|afternoon|evening|mat|ves|not)\b"
    trimester_pattern = r"\b\d+\s*(?:o|°)?\s*tri\b|\btrimestre\b"
    shift_hint = ""

    if any(token in normalized for token in [" matutino", " manha", " morning", " mat.", " mat "]):
        shift_hint = "morning"
    elif any(token in normalized for token in [" vespertino", " tarde", " afternoon", " ves.", " ves "]):
        shift_hint = "afternoon"
    elif any(token in normalized for token in [" noturno", " noite", " evening", " not.", " not "]):
        shift_hint = "evening"

    school_hint = ""
    school_match = re.search(r"escola\s*[:\-]?\s*([a-zA-Z0-9\- ]{3,80})", normalized)
    if school_match:
        school_raw = re.sub(shift_pattern, "", school_match.group(1))
        school_raw = re.sub(trimester_pattern, "", school_raw)
        school_raw = school_raw.replace(".", " ")
        school_raw = re.sub(r"\s+", " ", school_raw).strip(" -_")
        school_hint = school_raw.title()
    else:
        cleaned = re.sub(shift_pattern, "", normalized)
        cleaned = re.sub(trimester_pattern, "", cleaned)
        cleaned = cleaned.replace(".", " ")
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_")
        if cleaned:
            school_hint = cleaned.title()

    aliases = {
        "juv": "Juvenal",
        "juvenal": "Juvenal",
        "arap": "Arapongas",
        "arapongas": "Arapongas",
        "mulde": "Mulde",
        "anna": "Anna Alves",
        "anna alves": "Anna Alves",
        "ana alves": "Anna Alves",
        "tancredo": "Tancredo",
        "maria": "Maria Helena",
        "maria helena": "Maria Helena",
        "mair ahelena": "Maria Helena",
        "maira helena": "Maria Helena",
        "mairahelena": "Maria Helena",
    }

    school_key = _normalize_text(school_hint)
    school_key = re.sub(r"\s+", " ", school_key).strip()
    for key, canonical in aliases.items():
        if school_key == key or school_key.startswith(f"{key} "):
            school_hint = canonical
            break

    return school_hint, shift_hint


def _normalize_class_group_name(raw: str) -> str:
    value = _normalize_text(raw)
    value = value.replace("°", "").replace("º", "").strip()

    ordinal_map = {
        "sexto": "6",
        "setimo": "7",
        "sétimo": "7",
        "oitavo": "8",
        "nono": "9",
        "decimo": "10",
        "décimo": "10",
    }
    if value in ordinal_map:
        return f"{ordinal_map[value]}°"

    number_match = re.search(r"\b(\d{1,2})\b", value)
    if number_match:
        return f"{number_match.group(1)}°"

    return str(raw or "").strip()


def _extract_class_hint_from_text(text: str) -> str:
    normalized = _normalize_text(text)
    turma_match = re.search(r"turma\s*[:\-]?\s*([a-z0-9\- ]{2,20})", normalized)
    if turma_match:
        return _normalize_class_group_name(turma_match.group(1))

    ordinal_match = re.search(r"\b(sexto|setimo|s[eé]timo|oitavo|nono|decimo|d[eé]cimo|\d{1,2})\b", normalized)
    if ordinal_match:
        return _normalize_class_group_name(ordinal_match.group(1))

    return ""


def _looks_like_placeholder_student(name: str) -> bool:
    normalized = _normalize_text(name)
    return normalized in {
        "nome aluno",
        "nome",
        "aluno",
        "observacoes",
        "observacao",
        "ddmmyyyy",
    }


def _shift_suffix(shift: str) -> str:
    if shift == "morning":
        return "Matutino"
    if shift == "afternoon":
        return "Vespertino"
    if shift == "evening":
        return "Noturno"
    return "Turno"


def _canonical_key(raw_key: str) -> str:
    normalized = _normalize_key(raw_key)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in {_normalize_key(alias) for alias in aliases}:
            return canonical
    return normalized


def _normalize_row(raw_row: dict) -> dict:
    normalized = {}
    for key, value in raw_row.items():
        canonical = _canonical_key(str(key))
        normalized[canonical] = str(value).strip() if value is not None else ""
    return normalized


def _coerce_shift(raw_shift: str) -> str:
    value = _normalize_key(raw_shift)
    if value in {"morning", "manha", "matutino"}:
        return "morning"
    if value in {"afternoon", "tarde", "vespertino"}:
        return "afternoon"
    if value in {"evening", "noite", "noturno"}:
        return "evening"
    return "morning"


def _parse_csv_rows(file_obj) -> list[dict]:
    content = file_obj.read().decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(content))
    return [_normalize_row(row) for row in reader if row]


def _parse_xlsx_rows(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    rows: list[dict] = []

    def parse_rows_from_header(
        sheet,
        header_idx: int,
        start_col: int,
        end_col: int,
        sheet_school_hint: str,
        sheet_shift_hint: str,
        sheet_class_hint: str,
    ) -> list[dict]:
        header_cells = [sheet.cell(row=header_idx, column=col).value for col in range(start_col, end_col + 1)]
        headers = [str(item).strip() if item is not None else "" for item in header_cells]
        local_rows: list[dict] = []

        class_hint = ""
        school_hint = sheet_school_hint
        shift_hint = sheet_shift_hint
        current_class_hint = sheet_class_hint
        for lookback in range(max(1, header_idx - 4), header_idx):
            context_values = [sheet.cell(row=lookback, column=col).value for col in range(start_col, end_col + 1)]
            context_text = " ".join(str(v).strip() for v in context_values if v is not None and str(v).strip())
            norm = _normalize_text(context_text)

            class_match = re.search(r"turma\s*[:\-]?\s*([0-9]{1,2}\s*[a-zA-Z]|[a-zA-Z0-9\- ]{2,20})", norm)
            if class_match and not class_hint:
                class_hint = _normalize_class_group_name(class_match.group(1))

            school_match = re.search(r"escola\s*[:\-]?\s*([a-zA-Z0-9\- ]{3,60})", norm)
            if school_match and not school_hint:
                school_hint = school_match.group(1).strip().title()

            if not shift_hint:
                if any(token in norm for token in ["matutino", "manha", "morning"]):
                    shift_hint = "morning"
                elif any(token in norm for token in ["vespertino", "tarde", "afternoon"]):
                    shift_hint = "afternoon"
                elif any(token in norm for token in ["noturno", "noite", "evening"]):
                    shift_hint = "evening"

        if not school_hint:
            school_hint = str(sheet.title or "Escola").strip()
        if class_hint:
            current_class_hint = class_hint

        for row_idx in range(header_idx + 1, sheet.max_row + 1):
            # Encerramos bloco quando um novo cabeçalho da mesma seção aparece.
            if row_idx > header_idx + 1:
                maybe_header = _canonical_key(str(sheet.cell(row=row_idx, column=start_col).value or ""))
                if maybe_header == "student":
                    break

            values = [sheet.cell(row=row_idx, column=col).value for col in range(start_col, end_col + 1)]
            if not any(value is not None and str(value).strip() for value in values):
                continue

            section_text = _normalize_text(" ".join(str(v).strip() for v in values if v is not None and str(v).strip()))
            section_class_hint = _extract_class_hint_from_text(section_text)
            if section_class_hint:
                current_class_hint = section_class_hint
                continue

            raw = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
            normalized = _normalize_row(raw)

            if normalized.get("class_group"):
                current_class_hint = _normalize_class_group_name(normalized["class_group"])

            if current_class_hint and not normalized.get("class_group"):
                normalized["class_group"] = current_class_hint
            if school_hint and not normalized.get("school"):
                normalized["school"] = school_hint
            if shift_hint and not normalized.get("shift"):
                normalized["shift"] = shift_hint

            student_name = normalized.get("student", "")
            if not student_name or _looks_like_placeholder_student(student_name):
                continue

            normalized["student"] = str(student_name).strip().strip(".")

            if normalized.get("student") and not normalized.get("enrollment_code"):
                auto_code = _normalize_key(f"{normalized.get('class_group', 'turma')}_{normalized['student']}")
                normalized["enrollment_code"] = f"AUTO_{auto_code[:32]}"

            local_rows.append(normalized)

        return local_rows

    for sheet in workbook.worksheets:
        sheet_school_hint, sheet_shift_hint = _extract_school_and_shift_hint(str(sheet.title or ""))
        sheet_class_hint = _extract_class_hint_from_text(str(sheet.title or ""))
        parsed_standard_layout = False

        # Layout padrão: cabeçalho na primeira linha.
        iterator = sheet.iter_rows(values_only=True)
        first_row = next(iterator, None)
        if first_row:
            first_headers = [str(item).strip() if item is not None else "" for item in first_row]
            first_row_keys = {_canonical_key(item) for item in first_headers if item}
            if "student" in first_row_keys:
                parsed_standard_layout = True
                for values in iterator:
                    if not values or not any(value is not None and str(value).strip() for value in values):
                        continue
                    row = {first_headers[index]: values[index] for index in range(min(len(first_headers), len(values)))}
                    normalized = _normalize_row(row)
                    if sheet_school_hint and not normalized.get("school"):
                        normalized["school"] = sheet_school_hint
                    if sheet_shift_hint and not normalized.get("shift"):
                        normalized["shift"] = sheet_shift_hint
                    if normalized.get("class_group"):
                        normalized["class_group"] = _normalize_class_group_name(normalized.get("class_group", ""))
                    elif sheet_class_hint:
                        normalized["class_group"] = sheet_class_hint
                    student_name = normalized.get("student", "")
                    if not student_name or _looks_like_placeholder_student(student_name):
                        continue
                    normalized["student"] = str(student_name).strip().strip(".")
                    if normalized.get("student") and not normalized.get("enrollment_code"):
                        auto_code = _normalize_key(normalized["student"])
                        normalized["enrollment_code"] = f"AUTO_{auto_code[:32]}"
                    rows.append(normalized)

        # Layout em blocos: múltiplas turmas lado a lado, cabeçalhos no meio da planilha.
        if parsed_standard_layout:
            continue

        for row_idx in range(1, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
            if not any(value is not None and str(value).strip() for value in row_values):
                continue

            header_starts: list[int] = []
            for col_idx, value in enumerate(row_values, start=1):
                key = _canonical_key(str(value or ""))
                if key == "student":
                    header_starts.append(col_idx)

            if not header_starts:
                continue

            for index, start_col in enumerate(header_starts):
                end_col = header_starts[index + 1] - 1 if index + 1 < len(header_starts) else sheet.max_column
                parsed = parse_rows_from_header(
                    sheet,
                    row_idx,
                    start_col,
                    end_col,
                    sheet_school_hint,
                    sheet_shift_hint,
                    sheet_class_hint,
                )
                rows.extend(parsed)

    # Filtra apenas linhas com dados mínimos para evitar contagem de "puladas" artificialmente alta.
    rows = [row for row in rows if row.get("student") and row.get("class_group")]

    # Deduplicação por campos críticos para evitar duplicatas entre estratégias de parsing.
    unique: dict[tuple[str, str, str, str], dict] = {}
    for row in rows:
        key = (
            row.get("school", ""),
            row.get("class_group", ""),
            row.get("student", ""),
            row.get("enrollment_code", ""),
        )
        if any(key):
            unique[key] = row
    return list(unique.values())


def _parse_pdf_rows(file_obj) -> list[dict]:
    reader = PdfReader(file_obj)
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        lines.extend([line.strip() for line in text.splitlines() if line.strip()])

    delimiter = None
    for candidate in [";", "|", ",", "\t"]:
        if any(candidate in line for line in lines):
            delimiter = candidate
            break

    if not lines:
        return []

    # Caso 1: PDF com delimitador explícito (;|,|\t)
    if delimiter:
        header_line = next((line for line in lines if delimiter in line), "")
        if header_line:
            headers = [col.strip() for col in header_line.split(delimiter)]
            rows: list[dict] = []
            started = False
            for line in lines:
                if not started:
                    if line == header_line:
                        started = True
                    continue
                if delimiter not in line:
                    continue
                values = [col.strip() for col in line.split(delimiter)]
                row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
                rows.append(_normalize_row(row))
            return rows

    # Caso 2: PDF em tabela com colunas separadas por múltiplos espaços.
    header_idx = -1
    split_rows: list[list[str]] = []
    for idx, line in enumerate(lines):
        parts = [p.strip() for p in re.split(r"\s{2,}", line) if p.strip()]
        if not parts:
            continue
        split_rows.append(parts)
        canonical_parts = {_canonical_key(p) for p in parts}
        if "student" in canonical_parts:
            header_idx = idx

    if header_idx == -1:
        return []

    header_parts = [p.strip() for p in re.split(r"\s{2,}", lines[header_idx]) if p.strip()]
    parsed_rows: list[dict] = []
    for line in lines[header_idx + 1 :]:
        parts = [p.strip() for p in re.split(r"\s{2,}", line) if p.strip()]
        if not parts:
            continue
        row = {header_parts[index]: parts[index] if index < len(parts) else "" for index in range(len(header_parts))}
        parsed_rows.append(_normalize_row(row))

    return parsed_rows


def parse_enrollment_file_rows(file_obj) -> list[dict]:
    extension = Path(file_obj.name or "").suffix.lower()
    if extension not in SUPPORTED_IMPORT_EXTENSIONS:
        raise ValueError(f"Formato nao suportado: {extension}")

    file_obj.seek(0)
    if extension == ".csv":
        return _parse_csv_rows(file_obj)
    if extension == ".xlsx":
        return _parse_xlsx_rows(file_obj)
    return _parse_pdf_rows(file_obj)


def _collect_rows_from_files(files: list) -> tuple[list[dict], list[str]]:
    all_rows: list[dict] = []
    errors: list[str] = []

    for uploaded_file in files:
        try:
            all_rows.extend(parse_enrollment_file_rows(uploaded_file))
        except Exception as exc:  # pragma: no cover - defensive for malformed files
            errors.append(f"{uploaded_file.name}: {exc}")

    return all_rows, errors


def import_schools(files: list) -> dict:
    rows, errors = _collect_rows_from_files(files)
    processed_rows = 0
    skipped_rows = 0
    imported_schools = set()

    for row in rows:
        school_name = row.get("school", "")
        if not school_name:
            skipped_rows += 1
            continue

        school, _ = School.objects.update_or_create(
            name=school_name,
            defaults={
                "city": row.get("city") or "Nao informado",
                "state": (row.get("state") or "SP").upper()[:2],
            },
        )
        imported_schools.add(school.id)
        processed_rows += 1

    return {
        "processed_rows": processed_rows,
        "skipped_rows": skipped_rows,
        "schools_created_or_updated": len(imported_schools),
        "errors": errors,
    }


def import_class_groups(files: list) -> dict:
    rows, errors = _collect_rows_from_files(files)
    processed_rows = 0
    skipped_rows = 0
    imported_schools = set()
    imported_class_groups = set()

    for row in rows:
        school_name = row.get("school", "")
        class_group_name = row.get("class_group", "")

        if not school_name or not class_group_name:
            skipped_rows += 1
            continue

        school, _ = School.objects.get_or_create(
            name=school_name,
            defaults={
                "city": row.get("city") or "Nao informado",
                "state": (row.get("state") or "SP").upper()[:2],
            },
        )
        imported_schools.add(school.id)

        shift = _coerce_shift(row.get("shift", ""))
        normalized_class_name = _normalize_class_group_name(class_group_name)
        class_group = ClassGroup.objects.filter(school=school, name=normalized_class_name).first()

        if class_group and class_group.shift != shift:
            normalized_class_name = f"{normalized_class_name} ({_shift_suffix(shift)})"
            class_group = ClassGroup.objects.filter(school=school, name=normalized_class_name).first()

        if not class_group:
            class_group = ClassGroup.objects.create(
                school=school,
                name=normalized_class_name,
                grade_level=row.get("grade_level") or normalized_class_name,
                shift=shift,
            )
        else:
            changed = False
            next_grade = row.get("grade_level") or class_group.grade_level
            if class_group.grade_level != next_grade:
                class_group.grade_level = next_grade
                changed = True
            if class_group.shift != shift:
                class_group.shift = shift
                changed = True
            if changed:
                class_group.save(update_fields=["grade_level", "shift"])

        imported_class_groups.add(class_group.id)
        processed_rows += 1

    return {
        "processed_rows": processed_rows,
        "skipped_rows": skipped_rows,
        "schools_created_or_updated": len(imported_schools),
        "class_groups_created_or_updated": len(imported_class_groups),
        "errors": errors,
    }


def import_students(files: list) -> dict:
    rows, errors = _collect_rows_from_files(files)
    processed_rows = 0
    skipped_rows = 0
    imported_schools = set()
    imported_class_groups = set()
    imported_students = set()

    for row in rows:
        school_name = row.get("school", "")
        class_group_name = row.get("class_group", "")
        student_name = row.get("student", "")
        enrollment_code = row.get("enrollment_code", "")

        if not all([school_name, class_group_name, student_name]):
            skipped_rows += 1
            continue

        if not enrollment_code:
            enrollment_code = f"AUTO_{_normalize_key(student_name)[:32]}"

        school, _ = School.objects.get_or_create(
            name=school_name,
            defaults={
                "city": row.get("city") or "Nao informado",
                "state": (row.get("state") or "SP").upper()[:2],
            },
        )
        imported_schools.add(school.id)

        shift = _coerce_shift(row.get("shift", ""))
        normalized_class_name = _normalize_class_group_name(class_group_name)
        class_group = ClassGroup.objects.filter(school=school, name=normalized_class_name).first()
        if class_group and class_group.shift != shift:
            normalized_class_name = f"{normalized_class_name} ({_shift_suffix(shift)})"
            class_group = ClassGroup.objects.filter(school=school, name=normalized_class_name).first()

        if not class_group:
            class_group = ClassGroup.objects.create(
                school=school,
                name=normalized_class_name,
                grade_level=row.get("grade_level") or normalized_class_name,
                shift=shift,
            )

        imported_class_groups.add(class_group.id)

        student, _ = Student.objects.update_or_create(
            school=school,
            enrollment_code=enrollment_code,
            defaults={
                "class_group": class_group,
                "full_name": str(student_name).strip(),
            },
        )
        imported_students.add(student.id)
        processed_rows += 1

    return {
        "processed_rows": processed_rows,
        "skipped_rows": skipped_rows,
        "schools_created_or_updated": len(imported_schools),
        "class_groups_created_or_updated": len(imported_class_groups),
        "students_created_or_updated": len(imported_students),
        "errors": errors,
    }


def import_enrollments(files: list) -> dict:
    imported_schools = set()
    imported_class_groups = set()
    imported_students = set()
    processed_rows = 0
    skipped_rows = 0
    errors: list[str] = []
    seen_enrollments_by_class: dict[int, set[str]] = {}

    for uploaded_file in files:
        try:
            rows = parse_enrollment_file_rows(uploaded_file)
        except Exception as exc:  # pragma: no cover - defensive for malformed files
            errors.append(f"{uploaded_file.name}: {exc}")
            continue

        for row in rows:
            school_name = row.get("school", "")
            class_group_name = row.get("class_group", "")
            student_name = row.get("student", "")
            enrollment_code = row.get("enrollment_code", "")
            shift = _coerce_shift(row.get("shift", ""))

            if not all([school_name, class_group_name, student_name, enrollment_code]):
                skipped_rows += 1
                continue

            school, _ = School.objects.get_or_create(
                name=school_name,
                defaults={
                    "city": row.get("city") or "Nao informado",
                    "state": (row.get("state") or "SP").upper()[:2],
                },
            )
            imported_schools.add(school.id)

            normalized_class_name = _normalize_class_group_name(class_group_name)
            class_group = ClassGroup.objects.filter(school=school, name=normalized_class_name).first()

            if class_group and class_group.shift != shift:
                normalized_class_name = f"{normalized_class_name} ({_shift_suffix(shift)})"
                class_group = ClassGroup.objects.filter(school=school, name=normalized_class_name).first()

            if not class_group:
                class_group = ClassGroup.objects.create(
                    school=school,
                    name=normalized_class_name,
                    grade_level=row.get("grade_level") or normalized_class_name,
                    shift=shift,
                )

            imported_class_groups.add(class_group.id)
            seen_enrollments_by_class.setdefault(class_group.id, set()).add(enrollment_code)

            student, _ = Student.objects.update_or_create(
                school=school,
                enrollment_code=enrollment_code,
                defaults={
                    "class_group": class_group,
                    "full_name": student_name,
                },
            )
            imported_students.add(student.id)
            processed_rows += 1

    # Sincroniza turmas importadas removendo alunos antigos não presentes no arquivo atual.
    removed_students = 0
    for class_group_id, seen_codes in seen_enrollments_by_class.items():
        to_remove = Student.objects.filter(class_group_id=class_group_id).exclude(enrollment_code__in=seen_codes)
        removed_students += to_remove.count()
        to_remove.delete()

    return {
        "processed_rows": processed_rows,
        "skipped_rows": skipped_rows,
        "schools_created_or_updated": len(imported_schools),
        "class_groups_created_or_updated": len(imported_class_groups),
        "students_created_or_updated": len(imported_students),
        "students_removed": removed_students,
        "errors": errors,
    }


def easter_date(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def country_observances(year: int) -> list[dict]:
    easter = easter_date(year)
    return [
        {"name": "Carnaval (segunda)", "date": (easter - timedelta(days=48)).isoformat(), "type": "observance"},
        {"name": "Carnaval (terca)", "date": (easter - timedelta(days=47)).isoformat(), "type": "observance"},
        {"name": "Quarta-feira de Cinzas", "date": (easter - timedelta(days=46)).isoformat(), "type": "observance"},
        {"name": "Sexta-feira Santa", "date": (easter - timedelta(days=2)).isoformat(), "type": "observance"},
        {"name": "Corpus Christi", "date": (easter + timedelta(days=60)).isoformat(), "type": "observance"},
    ]


def country_dates(year: int) -> list[dict]:
    holidays = fetch_national_holidays(year)
    normalized_holidays = [
        {
            "name": holiday.get("name", "Feriado nacional"),
            "date": holiday.get("date"),
            "type": holiday.get("type") or "national_holiday",
            "source": "brasilapi",
        }
        for holiday in holidays
    ]

    observances = [
        {
            "name": item["name"],
            "date": item["date"],
            "type": item["type"],
            "source": "rule_based",
        }
        for item in country_observances(year)
    ]
    return sorted(normalized_holidays + observances, key=lambda item: item["date"])


def fetch_national_holidays(year: int) -> list[dict]:
    cached = NationalHolidayCache.objects.filter(year=year).first()
    if cached:
        return cached.raw_payload

    response = requests.get(BRASIL_API_HOLIDAYS.format(year=year), timeout=10)
    response.raise_for_status()
    payload = response.json()
    NationalHolidayCache.objects.update_or_create(year=year, defaults={"raw_payload": payload})
    return payload


def sync_national_holidays(calendar: AcademicCalendar) -> int:
    holidays = fetch_national_holidays(calendar.year)
    created = 0
    for holiday in holidays:
        _, was_created = CalendarEvent.objects.get_or_create(
            calendar=calendar,
            event_type=EventType.NATIONAL_HOLIDAY,
            title=holiday["name"],
            date=holiday["date"],
            defaults={"description": "Sincronizado via BrasilAPI"},
        )
        if was_created:
            created += 1
    return created


def class_group_teaching_weekdays(class_group: ClassGroup) -> set[int]:
    weekdays = set(
        WeeklyClassSlot.objects.filter(class_group=class_group)
        .values_list("weekday", flat=True)
        .distinct()
    )
    if weekdays:
        return weekdays
    return {
        Weekday.MONDAY,
        Weekday.TUESDAY,
        Weekday.WEDNESDAY,
        Weekday.THURSDAY,
        Weekday.FRIDAY,
    }


def consolidated_calendar_days(school: School, start: date, end: date) -> set[date]:
    return set(
        CalendarEvent.objects.filter(
            calendar__school=school,
            date__gte=start,
            date__lte=end,
        ).values_list("date", flat=True)
    )


def is_teaching_day(*, school: School, class_group: ClassGroup, day: date) -> bool:
    if day.weekday() in (Weekday.SATURDAY, Weekday.SUNDAY):
        return False

    event_exists = CalendarEvent.objects.filter(
        Q(calendar__school=school),
        Q(date=day),
        Q(class_group__isnull=True) | Q(class_group=class_group),
    ).exists()
    if event_exists:
        return False

    teaching_weekdays = class_group_teaching_weekdays(class_group)
    return day.weekday() in teaching_weekdays


def calculate_notification_date(*, school: School, class_group: ClassGroup, assessment_date: date, lead_days: int) -> date:
    current_day = assessment_date
    remaining = lead_days

    while remaining > 0:
        current_day -= timedelta(days=1)
        if is_teaching_day(school=school, class_group=class_group, day=current_day):
            remaining -= 1

    return current_day
