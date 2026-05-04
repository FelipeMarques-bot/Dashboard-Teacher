from datetime import date
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from rest_framework.test import APIClient
from openpyxl import Workbook

from .models import AcademicCalendar, CalendarEvent, ClassGroup, EventType, ShiftType, School, Teacher, Weekday, WeeklyClassSlot
from .services import calculate_notification_date


class NotificationDateServiceTests(TestCase):
    def setUp(self):
        self.school = School.objects.create(name="Escola Central", city="Sao Paulo", state="SP")
        self.class_group = ClassGroup.objects.create(
            school=self.school,
            name="7A",
            grade_level="7",
            shift=ShiftType.MORNING,
        )

        # Turma com aula de segunda a sexta.
        for weekday in [
            Weekday.MONDAY,
            Weekday.TUESDAY,
            Weekday.WEDNESDAY,
            Weekday.THURSDAY,
            Weekday.FRIDAY,
        ]:
            WeeklyClassSlot.objects.create(
                class_group=self.class_group,
                subject_id=self._get_subject_id(),
                weekday=weekday,
                slot_order=1,
            )

    def _get_subject_id(self):
        from .models import Subject

        subject, _ = Subject.objects.get_or_create(name="Matematica", school=self.school)
        return subject.id

    def test_ignores_weekends_and_holidays(self):
        calendar = AcademicCalendar.objects.create(
            school=self.school,
            year=2026,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
        )
        CalendarEvent.objects.create(
            calendar=calendar,
            event_type=EventType.MUNICIPAL_HOLIDAY,
            title="Feriado municipal",
            date=date(2026, 4, 20),
        )

        notification_date = calculate_notification_date(
            school=self.school,
            class_group=self.class_group,
            assessment_date=date(2026, 4, 23),
            lead_days=3,
        )

        self.assertEqual(notification_date, date(2026, 4, 17))


class ModelSmokeTests(TestCase):
    def test_teacher_profile_creation(self):
        school = School.objects.create(name="Escola B", city="Recife", state="PE")
        user = get_user_model().objects.create_user(username="prof1", password="abc12345")
        teacher = Teacher.objects.create(user=user, school=school)

        self.assertEqual(teacher.school.name, "Escola B")


class ImportAndCountryDatesApiTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username="tester", password="senha12345")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_import_enrollments_csv(self):
        csv_content = (
            "escola,cidade,estado,turma,serie,turno,aluno,matricula\n"
            "Escola Centro,Sao Paulo,SP,9A,9,Manha,Ana Silva,MAT001\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("alunos.csv", csv_content, content_type="text/csv")

        response = self.client.post("/api/import/enrollments/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["processed_rows"], 1)
        self.assertEqual(School.objects.count(), 1)
        self.assertEqual(ClassGroup.objects.count(), 1)
        from .models import Student

        self.assertEqual(Student.objects.count(), 1)

    def test_import_enrollments_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["escola", "cidade", "estado", "turma", "serie", "turno", "aluno", "matricula"])
        sheet.append(["Escola Sul", "Campinas", "SP", "8B", "8", "Tarde", "Bruno Lima", "MAT002"])

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        upload = SimpleUploadedFile(
            "alunos.xlsx",
            stream.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/import/enrollments/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["processed_rows"], 1)

    def test_import_enrollments_xlsx_uses_sheet_school_and_shift(self):
        workbook = Workbook()
        morning_sheet = workbook.active
        morning_sheet.title = "Escola Centro - Matutino"
        morning_sheet.append(["turma", "aluno", "matricula"])
        morning_sheet.append(["7A", "Ana Souza", "MAT100"])

        evening_sheet = workbook.create_sheet("Escola Norte Noturno")
        evening_sheet.append(["turma", "aluno", "matricula"])
        evening_sheet.append(["8B", "Bruno Reis", "MAT200"])

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        upload = SimpleUploadedFile(
            "abas.xlsx",
            stream.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/import/enrollments/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["processed_rows"], 2)

        class_morning = ClassGroup.objects.get(name="7A")
        class_evening = ClassGroup.objects.get(name="8B")
        self.assertEqual(class_morning.shift, ShiftType.MORNING)
        self.assertEqual(class_evening.shift, ShiftType.EVENING)
        self.assertEqual(class_morning.school.name, "Centro")
        self.assertEqual(class_evening.school.name, "Norte")

    def test_import_enrollments_xlsx_sheet_names_like_spreadsheet_tabs(self):
        workbook = Workbook()
        sheet_juv = workbook.active
        sheet_juv.title = "Juv.1°Tri"
        sheet_juv.append(["turma", "aluno", "matricula"])
        sheet_juv.append(["7A", "Ana Souza", "JUV001"])

        sheet_tancredo_mat = workbook.create_sheet("Tancredo mat. 1°Tri")
        sheet_tancredo_mat.append(["turma", "aluno", "matricula"])
        sheet_tancredo_mat.append(["7B", "Bruno Lima", "TANM001"])

        sheet_tancredo_ves = workbook.create_sheet("Tancredo ves. 1°Tri")
        sheet_tancredo_ves.append(["turma", "aluno", "matricula"])
        sheet_tancredo_ves.append(["7C", "Carla Mendes", "TANV001"])

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        upload = SimpleUploadedFile(
            "abas_trimestre.xlsx",
            stream.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/import/enrollments/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["processed_rows"], 3)

        class_mat = ClassGroup.objects.get(name="7B")
        class_ves = ClassGroup.objects.get(name="7C")
        class_juv = ClassGroup.objects.get(name="7A")

        self.assertEqual(class_mat.school.name, "Tancredo")
        self.assertEqual(class_ves.school.name, "Tancredo")
        self.assertEqual(class_juv.school.name, "Juvenal")
        self.assertEqual(class_mat.shift, ShiftType.MORNING)
        self.assertEqual(class_ves.shift, ShiftType.AFTERNOON)

    def test_import_enrollments_xlsx_real_school_tab_sequence(self):
        workbook = Workbook()
        tabs = [
            ("Juvenal", "6°", "JUV001", "Ana Silva"),
            ("Arapongas", "7°", "ARA001", "Bruno Lima"),
            ("Mulde", "8°", "MUL001", "Carla Rocha"),
            ("Anna Alves", "9°", "ANN001", "Daniel Souza"),
            ("Tancredo matutino", "6°", "TANM001", "Elisa Matos"),
            ("Tancredo vespertino", "6°", "TANV001", "Felipe Nunes"),
            ("Maria Helena matutino", "7°", "MHM001", "Gabriela Reis"),
            ("Mair aHelena vespertino", "7°", "MHV001", "Heitor Gomes"),
        ]

        first = workbook.active
        first.title = tabs[0][0]
        first.append(["Tr", "Nome aluno", "matricula"])
        first.append([tabs[0][1], tabs[0][3], tabs[0][2]])

        for title, turma, matricula, nome in tabs[1:]:
            sheet = workbook.create_sheet(title)
            sheet.append(["Tr", "Nome aluno", "matricula"])
            sheet.append([turma, nome, matricula])

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        upload = SimpleUploadedFile(
            "abas_reais.xlsx",
            stream.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/import/enrollments/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["processed_rows"], 8)

        self.assertTrue(School.objects.filter(name="Juvenal").exists())
        self.assertTrue(School.objects.filter(name="Arapongas").exists())
        self.assertTrue(School.objects.filter(name="Mulde").exists())
        self.assertTrue(School.objects.filter(name="Anna Alves").exists())
        self.assertTrue(School.objects.filter(name="Tancredo").exists())
        self.assertTrue(School.objects.filter(name="Maria Helena").exists())

        turma_tancredo_mat = ClassGroup.objects.get(school__name="Tancredo", name="6°")
        turma_tancredo_ves = ClassGroup.objects.get(school__name="Tancredo", name="6° (Vespertino)")
        self.assertEqual(turma_tancredo_mat.shift, ShiftType.MORNING)
        self.assertEqual(turma_tancredo_ves.shift, ShiftType.AFTERNOON)

    def test_import_enrollments_sync_removes_stale_students(self):
        first_csv = (
            "escola,cidade,estado,turma,serie,turno,aluno,matricula\n"
            "Escola Centro,Sao Paulo,SP,9A,9,Manha,Ana Silva,MAT001\n"
            "Escola Centro,Sao Paulo,SP,9A,9,Manha,Bruno Dias,MAT002\n"
        ).encode("utf-8")
        first_upload = SimpleUploadedFile("primeiro.csv", first_csv, content_type="text/csv")

        first_response = self.client.post("/api/import/enrollments/", {"files": [first_upload]}, format="multipart")
        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(first_response.data["processed_rows"], 2)

        second_csv = (
            "escola,cidade,estado,turma,serie,turno,aluno,matricula\n"
            "Escola Centro,Sao Paulo,SP,9A,9,Manha,Ana Silva,MAT001\n"
        ).encode("utf-8")
        second_upload = SimpleUploadedFile("segundo.csv", second_csv, content_type="text/csv")

        second_response = self.client.post("/api/import/enrollments/", {"files": [second_upload]}, format="multipart")
        self.assertEqual(second_response.status_code, 200)
        self.assertEqual(second_response.data["students_removed"], 1)

        from .models import Student

        self.assertEqual(Student.objects.filter(class_group__name="9A").count(), 1)

    def test_country_dates_endpoint(self):
        response = self.client.get("/api/calendar/country-dates/?year=2026")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["year"], 2026)
        self.assertTrue(len(response.data["dates"]) > 0)

    def test_import_schools_endpoint(self):
        csv_content = (
            "escola,cidade,estado\n"
            "Escola Alfa,Londrina,PR\n"
            "Escola Beta,Curitiba,PR\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("escolas.csv", csv_content, content_type="text/csv")

        response = self.client.post("/api/import/schools/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["schools_created_or_updated"], 2)

    def test_import_class_groups_endpoint(self):
        csv_content = (
            "escola,cidade,estado,turma,serie,turno\n"
            "Escola Alfa,Londrina,PR,6A,6,Manha\n"
            "Escola Alfa,Londrina,PR,6A,6,Tarde\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("turmas.csv", csv_content, content_type="text/csv")

        response = self.client.post("/api/import/class-groups/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["class_groups_created_or_updated"] >= 2)

    def test_import_students_endpoint(self):
        csv_content = (
            "escola,cidade,estado,turma,serie,turno,aluno,matricula\n"
            "Escola Alfa,Londrina,PR,6A,6,Manha,Ana Silva,ALU001\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("alunos_segmentado.csv", csv_content, content_type="text/csv")

        response = self.client.post("/api/import/students/", {"files": [upload]}, format="multipart")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["students_created_or_updated"], 1)


class GoogleLoginApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()

    @patch("apps.school.auth_views._initialize_firebase", return_value=None)
    @patch(
        "apps.school.auth_views.firebase_auth.verify_id_token",
        return_value={
            "email": "google.user@example.com",
            "name": "Google User",
            "given_name": "Google",
            "family_name": "User",
        },
    )
    def test_google_login_returns_jwt_and_creates_user(self, _verify_token, _init_firebase):
        response = self.client.post("/api/auth/google/", {"id_token": "fake-token"}, format="json")

        self.assertEqual(response.status_code, 200)
        self.assertIn("access", response.data)
        self.assertIn("refresh", response.data)
        self.assertEqual(response.data["user"]["email"], "google.user@example.com")
        self.assertEqual(get_user_model().objects.filter(email="google.user@example.com").count(), 1)
